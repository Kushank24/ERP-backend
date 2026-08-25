from __future__ import annotations

import csv
import io
from collections import OrderedDict
from typing import List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from pydantic import BaseModel, Field
from sqlalchemy import text
from sqlalchemy.orm import Session

from ..boq_math import total_quantity_consumed
from ..db import get_db
from ..deps import get_current_user

router = APIRouter(prefix="/products", tags=["products"])


class BoqLineIn(BaseModel):
    name: str = Field(min_length=1)
    section_size: float = Field(ge=0)
    units: str = Field(min_length=1)
    quantity: float = Field(gt=0)


class ProductCreate(BaseModel):
    name: str = Field(min_length=1)
    product_code: Optional[str] = None
    description: Optional[str] = None
    category: Optional[str] = None
    default_unit_price: float = Field(ge=0, default=0)
    boq_lines: List[BoqLineIn] = []


class ProductPatch(BaseModel):
    name: Optional[str] = None
    product_code: Optional[str] = None
    description: Optional[str] = None
    category: Optional[str] = None
    default_unit_price: Optional[float] = Field(default=None, ge=0)


class BoqLineAdd(BaseModel):
    name: str = Field(min_length=1)
    section_size: float = Field(ge=0)
    units: str = Field(min_length=1)
    quantity: float = Field(gt=0)


class BoqLinePatch(BaseModel):
    name: Optional[str] = Field(default=None, min_length=1)
    section_size: Optional[float] = Field(default=None, ge=0)
    units: Optional[str] = Field(default=None, min_length=1)
    quantity: Optional[float] = Field(default=None, gt=0)


def _load_boq(db: Session, product_id: int) -> list[dict]:
    rows = db.execute(
        text(
            """
            SELECT b.id, b.name, b.section_size, b.units, b.quantity, b.total_quantity_consumed
            FROM bill_of_quantities b
            JOIN product_bill_of_quantity_relations r ON r.bill_of_quantity_id = b.id
            WHERE r.product_id = :pid
            ORDER BY b.name
            """
        ),
        {"pid": product_id},
    ).mappings().all()
    return [dict(x) for x in rows]


def _product_dict(db: Session, product_id: int) -> dict:
    r = db.execute(
        text(
            """
            SELECT id, name, product_code, description, category, default_unit_price, created_at, updated_at
            FROM products WHERE id = :id
            """
        ),
        {"id": product_id},
    ).mappings().first()
    if not r:
        raise HTTPException(404, "Product not found")
    d = dict(r)
    d["bill_of_quantities"] = _load_boq(db, product_id)
    return d


@router.get("/categories")
def list_categories(db: Session = Depends(get_db), user: dict = Depends(get_current_user)):
    _ = user
    rows = db.execute(
        text("SELECT DISTINCT category FROM products ORDER BY category NULLS LAST")
    ).scalars().all()
    return [r or "" for r in rows]


@router.get("")
def list_products(
    search: Optional[str] = Query(default=None),
    category: Optional[str] = Query(default=None),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=100, ge=1, le=500),
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    _ = user
    conditions: list[str] = []
    params: dict = {}

    if search:
        conditions.append(
            "(p.name ILIKE :search OR p.product_code ILIKE :search OR p.category ILIKE :search)"
        )
        params["search"] = f"%{search.strip()}%"

    if category is not None:
        if category == "":
            conditions.append("(p.category IS NULL OR p.category = '')")
        else:
            conditions.append("p.category = :category")
            params["category"] = category

    where = f"WHERE {' AND '.join(conditions)}" if conditions else ""

    total: int = db.execute(
        text(f"SELECT COUNT(*) FROM products p {where}"), params
    ).scalar_one()

    params["limit"] = page_size
    params["offset"] = (page - 1) * page_size

    rows = db.execute(
        text(
            f"""
            SELECT p.id, p.name, p.product_code, p.description, p.category,
                   p.default_unit_price, p.created_at, p.updated_at,
                   COUNT(r.bill_of_quantity_id)::int AS boq_count
            FROM products p
            LEFT JOIN product_bill_of_quantity_relations r ON r.product_id = p.id
            {where}
            GROUP BY p.id, p.name, p.product_code, p.description, p.category,
                     p.default_unit_price, p.created_at, p.updated_at
            ORDER BY p.name
            LIMIT :limit OFFSET :offset
            """
        ),
        params,
    ).mappings().all()

    return {
        "items": [dict(r) for r in rows],
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@router.get("/{product_id}")
def get_product(product_id: int, db: Session = Depends(get_db), user: dict = Depends(get_current_user)):
    _ = user
    return _product_dict(db, product_id)


@router.post("", status_code=201)
def create_product(
    body: ProductCreate,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    _ = user
    pr = db.execute(
        text(
            """
            INSERT INTO products (name, product_code, description, category, default_unit_price)
            VALUES (:name, :code, :desc, :cat, :price)
            RETURNING id
            """
        ),
        {
            "name": body.name.strip(),
            "code": body.product_code,
            "desc": body.description,
            "cat": body.category,
            "price": body.default_unit_price,
        },
    ).first()
    pid = pr[0]

    seen: set[tuple[str, float]] = set()
    for line in body.boq_lines:
        key = (line.name.strip(), float(line.section_size))
        if key in seen:
            db.rollback()
            raise HTTPException(400, f"Duplicate BOQ line for {line.name} / {line.section_size}")
        seen.add(key)
        tq = total_quantity_consumed(line.section_size, line.quantity)
        boq = db.execute(
            text(
                """
                INSERT INTO bill_of_quantities (name, section_size, units, quantity, total_quantity_consumed)
                VALUES (:name, :ss, :units, :qty, :tq)
                RETURNING id
                """
            ),
            {
                "name": line.name.strip(),
                "ss": line.section_size,
                "units": line.units.strip(),
                "qty": line.quantity,
                "tq": tq,
            },
        ).first()
        bid = boq[0]
        db.execute(
            text(
                """
                INSERT INTO product_bill_of_quantity_relations (product_id, bill_of_quantity_id)
                VALUES (:pid, :bid)
                """
            ),
            {"pid": pid, "bid": bid},
        )

    db.commit()
    return _product_dict(db, pid)


@router.patch("/{product_id}")
def patch_product(
    product_id: int,
    body: ProductPatch,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    _ = user
    row = db.execute(text("SELECT id FROM products WHERE id = :id"), {"id": product_id}).first()
    if not row:
        raise HTTPException(404, "Product not found")
    fields = body.model_dump(exclude_unset=True)
    if fields:
        sets = []
        params: dict = {"id": product_id}
        for k, v in fields.items():
            sets.append(f"{k} = :{k}")
            params[k] = v
        sets.append("updated_at = now()")
        db.execute(text(f"UPDATE products SET {', '.join(sets)} WHERE id = :id"), params)
        db.commit()
    return _product_dict(db, product_id)


@router.post("/{product_id}/boq", status_code=201)
def add_boq_line(
    product_id: int,
    body: BoqLineAdd,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    _ = user
    row = db.execute(text("SELECT id FROM products WHERE id = :id"), {"id": product_id}).first()
    if not row:
        raise HTTPException(404, "Product not found")
    tq = total_quantity_consumed(body.section_size, body.quantity)
    boq = db.execute(
        text(
            """
            INSERT INTO bill_of_quantities (name, section_size, units, quantity, total_quantity_consumed)
            VALUES (:name, :ss, :units, :qty, :tq)
            RETURNING id
            """
        ),
        {
            "name": body.name.strip(),
            "ss": body.section_size,
            "units": body.units.strip(),
            "qty": body.quantity,
            "tq": tq,
        },
    ).first()
    bid = boq[0]
    db.execute(
        text(
            """
            INSERT INTO product_bill_of_quantity_relations (product_id, bill_of_quantity_id)
            VALUES (:pid, :bid)
            """
        ),
        {"pid": product_id, "bid": bid},
    )
    db.commit()
    return _product_dict(db, product_id)


@router.patch("/{product_id}/boq/{boq_id}")
def patch_boq_line(
    product_id: int,
    boq_id: int,
    body: BoqLinePatch,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    _ = user
    rel = db.execute(
        text(
            """
            SELECT 1 FROM product_bill_of_quantity_relations
            WHERE product_id = :pid AND bill_of_quantity_id = :bid
            """
        ),
        {"pid": product_id, "bid": boq_id},
    ).first()
    if not rel:
        raise HTTPException(404, "BOQ relation not found")
    fields = body.model_dump(exclude_unset=True)
    if fields:
        current = db.execute(
            text("SELECT section_size, quantity FROM bill_of_quantities WHERE id = :id"),
            {"id": boq_id},
        ).mappings().first()
        ss = fields.get("section_size", float(current["section_size"]))
        qty = fields.get("quantity", float(current["quantity"]))
        fields["total_quantity_consumed"] = total_quantity_consumed(ss, qty)
        sets = [f"{k} = :{k}" for k in fields]
        params: dict = {"id": boq_id, **fields}
        db.execute(text(f"UPDATE bill_of_quantities SET {', '.join(sets)} WHERE id = :id"), params)
        db.commit()
    return _product_dict(db, product_id)


@router.post("/bulk-upload", status_code=201)
async def bulk_upload_products(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    _ = user
    content = await file.read()
    filename = (file.filename or "").lower()

    rows: list[dict] = []
    if filename.endswith(".csv"):
        text_io = io.StringIO(content.decode("utf-8-sig"))
        rows = list(csv.DictReader(text_io))
    elif filename.endswith((".xlsx", ".xls")):
        try:
            import openpyxl
        except ImportError:
            raise HTTPException(400, "openpyxl not installed; please upload a CSV file instead")
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, [str(v) if v is not None else "" for v in row])))
    else:
        raise HTTPException(400, "Unsupported file type; upload a .csv or .xlsx file")

    def _norm(d: dict) -> dict:
        return {k.strip().lower().replace(" ", "_"): (v.strip() if isinstance(v, str) else str(v or "")) for k, v in d.items()}

    rows = [_norm(r) for r in rows if any(v for v in r.values())]

    products_map: dict = OrderedDict()
    parse_errors: list[str] = []

    for i, row in enumerate(rows, start=2):
        prod_name = row.get("name") or row.get("product_name", "")
        if not prod_name:
            parse_errors.append(f"Row {i}: missing product name")
            continue
        if prod_name not in products_map:
            try:
                price = float(row.get("default_unit_price") or 0)
            except (ValueError, TypeError):
                price = 0.0
            products_map[prod_name] = {
                "name": prod_name,
                "product_code": row.get("product_code") or None,
                "category": row.get("category") or None,
                "default_unit_price": price,
                "boq_lines": [],
            }
        material = row.get("material_name") or row.get("material", "")
        if material:
            try:
                products_map[prod_name]["boq_lines"].append({
                    "name": material,
                    "section_size": float(row.get("section_size") or 0),
                    "units": row.get("units") or "Nos",
                    "quantity": float(row.get("quantity") or 1),
                })
            except (ValueError, TypeError) as exc:
                parse_errors.append(f"Row {i}: invalid numeric value — {exc}")

    created_names: list[str] = []
    skipped_names: list[str] = []

    for prod_name, pdata in products_map.items():
        existing = db.execute(
            text("SELECT id FROM products WHERE name = :name"),
            {"name": pdata["name"]},
        ).first()
        if existing:
            skipped_names.append(prod_name)
            continue
        pr = db.execute(
            text(
                """
                INSERT INTO products (name, product_code, description, category, default_unit_price)
                VALUES (:name, :code, NULL, :cat, :price)
                RETURNING id
                """
            ),
            {"name": pdata["name"], "code": pdata["product_code"], "cat": pdata["category"], "price": pdata["default_unit_price"]},
        ).first()
        pid = pr[0]
        seen: set[tuple] = set()
        for line in pdata["boq_lines"]:
            key = (line["name"], float(line["section_size"]))
            if key in seen:
                continue
            seen.add(key)
            tq = total_quantity_consumed(line["section_size"], line["quantity"])
            boq = db.execute(
                text(
                    """
                    INSERT INTO bill_of_quantities (name, section_size, units, quantity, total_quantity_consumed)
                    VALUES (:name, :ss, :units, :qty, :tq)
                    RETURNING id
                    """
                ),
                {"name": line["name"], "ss": line["section_size"], "units": line["units"], "qty": line["quantity"], "tq": tq},
            ).first()
            db.execute(
                text("INSERT INTO product_bill_of_quantity_relations (product_id, bill_of_quantity_id) VALUES (:pid, :bid)"),
                {"pid": pid, "bid": boq[0]},
            )
        db.commit()
        created_names.append(prod_name)

    return {
        "created": len(created_names),
        "skipped": len(skipped_names),
        "created_names": created_names,
        "skipped_names": skipped_names,
        "errors": parse_errors,
    }


@router.delete("/{product_id}/boq/{boq_id}", status_code=200)
def remove_boq_line(
    product_id: int,
    boq_id: int,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    _ = user
    rel = db.execute(
        text(
            """
            SELECT 1 FROM product_bill_of_quantity_relations
            WHERE product_id = :pid AND bill_of_quantity_id = :bid
            """
        ),
        {"pid": product_id, "bid": boq_id},
    ).first()
    if not rel:
        raise HTTPException(404, "BOQ relation not found")
    db.execute(
        text(
            """
            DELETE FROM product_bill_of_quantity_relations
            WHERE product_id = :pid AND bill_of_quantity_id = :bid
            """
        ),
        {"pid": product_id, "bid": boq_id},
    )
    db.execute(
        text("DELETE FROM bill_of_quantities WHERE id = :id"),
        {"id": boq_id},
    )
    db.commit()
    return _product_dict(db, product_id)
